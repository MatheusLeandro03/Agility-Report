import os
import time
import subprocess
from datetime import datetime
from dateutil.relativedelta import relativedelta
import platform

def limpar_tela():
    # Função para limpar a tela em diferentes sistemas operacionais
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

limpar_tela()

# Instalação dos Pacotes
while True: 
    opcaoDownload = input ("Deseja instalar/atualizar os pacotes? (Y/N): ")
    opcaoDownload = opcaoDownload.lower()
    if opcaoDownload == "y":
        print ("As bibliotecas estão sendo instaladas/atualizadas para aprimorar o funcionamento do programa")
        time.sleep(1)
        subprocess.run(["pip", "install", "pandas"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.run(["pip", "install", "openpyxl"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        limpar_tela()
        print ("Pacotes instalados/atualizados! Iniciaremos o programa!!!")
        time.sleep(1)
        break
    elif opcaoDownload == "n":
        limpar_tela()
        print ("Continuaremos a execução do programa!")
        break
    else:
        print ("Opção inválida!")
        time.sleep(2)

# Início do Programa
import pandas as pd
from openpyxl import load_workbook

ExcelCliente1="Gráficos CLIENTE1.xlsx"
ExcelCliente2="Gráficos CLIENTE2.xlsx"
ExcelCliente3="Gráficos CLIENTE3.xlsx"

agrupamentos = {
    'http://Cliente1.com.br': ['http://cliente1.com.br/'],
    'https://testeCliente1.com.br': ['https://testeCliente1.com.br']
}
origensCliente1 = {
    'http://Cliente1.com.br': 0,
    'https://testeCliente1.com.br': 0
}
origensCliente2 = {
        'http://Cliente2.com.br': 0,
        'https://testeCliente2.com.br': 0
}
origensCliente3 = {
        'http://Cliente3.com.br': 0,
        'https://testeCliente3.com.br': 0
}

celulaTotalAPIsCliente2=279
celulaTotalAPIs = "C269"

#Linhas dos Sistemas Operacionais e Origens (Sites) dos Clientes (Cookies)
linhainicialSoSCliente1 = 136; linhaInicialSitesCliente1 = 200
linhainicialSoSCliente3 = 193; linhaInicialSitesCliente3= 173
linhainicialSoSCliente2 = 150; linhaInicialSitesCliente2 = 219

#Linhas das DSRs dos Clientes
linhainicialDSRsCliente1 = 27; celulaTotalCliente1 = "C19"
linhainicialDSRsCliente2 = 28; celulaTotalCliente2 = "C21"; celulaNaoVenda = "C55"

#Linhas dos Status de Cookies dos Clientes
celulaConcedidosCliente1 = "C129"; celulaRetiradosCliente1 = "D129"
celulaIgnoradosCliente1 = "E129"; celulaRecusadosCliente1 = "F129"
celulaStatusConcedidosCliente1 = "C160"; celulaStatusRetiradosCliente1 = "C162"
celulaStatusIgnoradosCliente1 = "C161"; celulaStatusRecusadosCliente1 = "C163"
 
celulaConcedidosCliente2 = "C178"; celulaIgnoradosCliente2 = "C179"
celulaRetiradosCliente2 = "C180"; celulaRecusadosCliente2 = "C181" 
celulaStatusConcedidosCliente2 = "C144"; celulaStatusIgnoradosCliente2 = "D144"
celulaStatusRetiradosCliente2 = "F144" ; celulaStatusRecusadosCliente2 = "E144"
celulaStatusConcedidosCliente2Ambiente = "C196"; celulaStatusIgnoradosCliente2Ambiente = "C197"
celulaStatusRetiradosCliente2Ambiente = "C198"; celulaStatusRecusadosCliente2Ambiente = "C199"

celulaConcedidosCliente3="C163"; celulaRecusadosCliente3="D163"; celulaRetiradosCliente3="E163"
celulaStatusConcedidosCliente3="C131"; celulaStatusRecusadosCliente3="C132"; celulaStatusRetiradosCliente3="C133"
celulaStatusIgnoradosCliente3="I5"; celulaIgnoradosCliente3="I6"

ConcedidoEssencialCliente1 = "C216"; RecusadoEssencialCliente1 = "D216"; RetiradoEssencialCliente1 = "E216"
ConcedidoPublicidadeCliente1 = "C217"; RecusadoPublicidadeCliente1 = "D217"; RetiradoPublicidadeCliente1 = "E217"
ConcedidoAnaliseEPersonalizacaoCliente1 = "C218"; RecusadoAnaliseEPersonalizacaoCliente1 = "D218"; 
RetiradoAnaliseEPersonalizacaoCliente1 = "E218"; ConcedidoDesempenhoEFuncionalidadeCliente1 = "C219"
RecusadoDesempenhoEFuncionalidadeCliente1 = "D219"; RetiradoDesempenhoEFuncionalidadeCliente1 = "E219"

ConcedidoEssencialCliente2 = "C235"; RecusadoEssencialCliente2 = "D235"; RetiradoEssencialCliente2 = "E235"
ConcedidoPublicidadeCliente2 = "C236"; RecusadoPublicidadeCliente2 = "D236"; RetiradoPublicidadeCliente2 = "E236"
ConcedidoAnaliseEPersonalizacaoCliente2 = "C237"; RecusadoAnaliseEPersonalizacaoCliente2 = "D237"; RetiradoAnaliseEPersonalizacaoCliente2 = "E237"
ConcedidoDesempenhoEFuncionalidadeCliente2 = "C238"; RecusadoDesempenhoEFuncionalidadeCliente2 = "D238"; RetiradoDesempenhoEFuncionalidadeCliente2 = "E238"

ConcedidoEssencialCliente3="C214";RecusadoEssencialCliente3="D214";RetiradoEssencialCliente3="E214"
ConcedidoPublicidadeCliente3="C215";RecusadoPublicidadeCliente3="D215";RetiradoPublicidadeCliente3="E215"
ConcedidoAnaliseEPersonalizacaoCliente3="C216";RecusadoAnaliseEPersonalizacaoCliente3="D216";RetiradoAnaliseEPersonalizacaoCliente3="E216"
ConcedidoDesempenhoEFuncionalidadeCliente3="C217";RecusadoDesempenhoEFuncionalidadeCliente3="D217";RetiradoDesempenhoEFuncionalidadeCliente3="E217"

linhaInicialConsentimentoUniversalCliente3=65
ConsentimentoUniversalConcedidoCliente2 = "C93"; ConsentimentoUniversalRetiradoCliente2="D93";ConsentimentoUniversalRecusadoCliente2="E93"
ConsentimentoUniversalConcedidoCliente3 = "C50"; ConsentimentoUniversalRecusadoCliente3 ="C51"; ConsentimentoUniversalRetiradoCliente3="C52"

NotificacoesConcedidoCliente2 ="C110";NotificacoesRetiradoCliente2 ="D110";NotificacoesRecusadoCliente2 ="E110"
SMSConcedidoCliente2 ="C111";SMSRetiradoCliente2 ="D111";SMSRecusadoCliente2 ="E111"
EmailConcedidoCliente2 ="C112";EmailRetiradoCliente2 ="D112";EmailRecusadoCliente2 ="E112"
TelefoneConcedidoCliente2 ="C113";TelefoneRetiradoCliente2 ="D113";TelefoneRecusadoCliente2 ="E113"
WhatsappConcedidoCliente2 ="C114";WhatsappRetiradoCliente2 ="D114";WhatsappRecusadoCliente2 ="E114"
OfertasConcedidoCliente2 ="C115";OfertasRetiradoCliente2 ="D115";OfertasRecusadoCliente2 ="E115"

TransmissaoAutorizadoCliente3="C70";TransmissaoRejeitadoCliente3="D70";TransmissaoAguardandoCliente3="E70"
RecepcaoAutorizadoCliente3="C69";RecepcaoRejeitadoCliente3="D69";RecepcaoAguardandoCliente3="E69"
somaAutorizadosCliente3="C91";somaRejeitadosCliente3="C92";somaAguardandoCliente3="C93"
somaAutorizadosHistoricoCliente3="C123";somaRejeitadosHistoricoCliente3="D123";somaAguardandoHistoricoCliente3="E123"
# Preparações Para Leitura
def leitura():
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    chunk_size = 1000000
    return chunk_size

# Salvar as alterações no arquivo Excel
def salvarExcel(graficoCliente,wb):
    wb.save(graficoCliente)
    for _ in range(1000000):
        pass
    time.sleep(5)

# Abre o arquivo gráficos
def CarregarExcel(graficoCliente):
    wb = load_workbook(graficoCliente)
    ws = wb.active
    return ws,wb
limpar_tela()
def tempoDecorrido(inicio):
    # Fim do contador de tempo
    fim = time.time()
    
    limpar_tela()

    # Cálculo do tempo decorrido
    tempo_decorrido = fim - inicio

    minutos_decorridos = int(tempo_decorrido // 60)
    segundos_decorridos = int(tempo_decorrido % 60)

    # Imprime o tempo decorrido
    print("\n------------------------------------------\nTempo decorrido: {} min e {} seg".format(minutos_decorridos, segundos_decorridos))
    time.sleep(5)
# Função para Ler Arquivos das DSRs
def leituraArquivosDSRs():
    while True: 
        limpar_tela()
        arquivo_DSR = input("Cole somente o nome do arquivo csv das \033[1mDSRs\033[0m: ")
        try:
            leituraDSRs = pd.read_csv(arquivo_DSR+'.csv', encoding="latin-1", sep=';')
            return leituraDSRs
            break
        except FileNotFoundError:
            print(f"O arquivo '{arquivo_DSR}' não foi encontrado. Certifique-se de que o nome do arquivo CSV esteja correto e sem sua extensão .csv!")
            time.sleep(3)
    limpar_tela()
# Função para Ler Arquivos dos Consentimentos Universais
def leituraArquivosConsentimentoUniversal():
    while True:
        limpar_tela()
        chunk_size=leitura()
        global arquivo_CosentimentoUniversal
        arquivo_ConsentimentoUniversal = input ("Informe somente o nome do arquivo csv dos Consentimentos Universais: ")
        try:
            global leituraConsentimentoUniversal 
            leituraConsentimentoUniversal = pd.read_csv(arquivo_ConsentimentoUniversal+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
            return leituraConsentimentoUniversal, arquivo_ConsentimentoUniversal
            break
        except FileNotFoundError:
            print(f"O arquivo '{arquivo_ConsentimentoUniversal}' não foi encontrado. Certifique-se de que o nome do arquivo CSV esteja correto e sem sua extensão .csv!")
            time.sleep(3)
    limpar_tela()
# Função para ler Arquivos de Cookies
def leituraArquivosCookies():

    while True:
        limpar_tela()
        chunk_size = leitura()
        global arquivo_Cookies
        arquivo_Cookies = input ("Informe somente o nome do arquivo csv dos Cookies: ")
        try:
            global leituraCookies 
            leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
            return leituraCookies, arquivo_Cookies
            break
        except FileNotFoundError:
            print(f"O arquivo '{arquivo_Cookies}' não foi encontrado. Certifique-se de que o nome do arquivo CSV esteja correto e sem sua extensão .csv!")
            time.sleep(3)
    limpar_tela()
# Função para ler Arquivos de Cookies Cliente 2 Segundo Ambiente
def leituraArquivosCookiesCliente2Ambiente():
    while True:
        chunk_size = leitura()
        limpar_tela()
        print("------ Cliente2 Segundo Ambiente ------")
        global arquivo_CookiesCliente2Ambiente
        arquivo_CookiesCliente2Ambiente = input ("Informe somente o nome do arquivo csv dos Cookies: ")
        try:
            global leituraCookiesCliente2Ambiente
            leituraCookiesCliente2Ambiente = pd.read_csv(arquivo_CookiesCliente2Ambiente+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
            return leituraCookiesCliente2Ambiente, arquivo_CookiesCliente2Ambiente
            break
        except FileNotFoundError:
            print(f"O arquivo '{arquivo_CookiesCliente2Ambiente}' não foi encontrado. Certifique-se de que o nome do arquivo CSV esteja correto e sem sua extensão .csv!")
            time.sleep(3)
    limpar_tela()    
# Função para ler Arquivos de APIs
def leituraArquivoAPIs():
    while True:
        chunk_size=leitura()
        limpar_tela()
        arquivo_APIs = input("Informe somente o nome do arquivo csv do uso de APIs: ")
        try:
            # Lê o arquivo em chunks e concatena em um DataFrame
            chunks = pd.read_csv(arquivo_APIs + '.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
            leituraAPIs = pd.concat(chunks, ignore_index=True)
            return leituraAPIs
        except FileNotFoundError:
            print(f"O arquivo '{arquivo_APIs}' não foi encontrado. Certifique-se de que o nome do arquivo CSV esteja correto e sem sua extensão .csv!")
            time.sleep(3)
    limpar_tela()
#Função para Filtrar as DSRs
def FiltragemDSRs(leituraDSRs, graficoCliente, opcao, linhainicialDSRs, celulaTotal, celulaNaoVenda):
    ws, wb = CarregarExcel(graficoCliente)    
    limpar_tela()
    print(leituraDSRs.columns)
    # Aplicar o filtro DSRs
    try:
        filtroAcesso = (leituraDSRs['Tipo'] == 'Acesso') & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroAcesso = (leituraDSRs['Tipo'] == 'Acesso') & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroEliminacao = (leituraDSRs['Tipo'] == 'Eliminacao') & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroDivSubProc = (leituraDSRs['Tipo'] == "Divulgacao de subprocessadores") & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroObjecao = (leituraDSRs['Tipo'] == 'Objecao') & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroPortabilidade = (leituraDSRs['Tipo'] == 'Portabilidade') & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroRetificacao = (leituraDSRs['Tipo'] == 'Retificacao') & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroConfirmacao = (leituraDSRs['Tipo'] == 'Confirmacao') & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroNaoVenda = (leituraDSRs['Tipo'] == "Não venda nem compartilhe") & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroProcessoRestrito = (leituraDSRs['Tipo'] == "Processo Restrito") & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
        filtroRevTomadaDecisao = (leituraDSRs['Tipo'] == "Revisar a Tomada de Decisãa Individual Automatizada") & (~leituraDSRs['ï»¿Nome do assunto'].str.contains('teste', case=False, na=False))
    except KeyError as e:
        # Aplicando filtros diretamente com str.contains para todos os tipos de DSR
        filtroAcesso = (leituraDSRs['Type'] == 'Access') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))
        filtroEliminacao = (leituraDSRs['Type'] == 'Erasure') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))
        filtroDivSubProc = (leituraDSRs['Type'] == 'Disclosure of Subprocessors') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))
        filtroObjecao = (leituraDSRs['Type'] == 'Objection') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))
        filtroPortabilidade = (leituraDSRs['Type'] == 'Port') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))
        filtroRetificacao = (leituraDSRs['Type'] == 'Rectification') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))
        filtroConfirmacao = (leituraDSRs['Type'] == 'Confirmation') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))
        filtroNaoVenda = (leituraDSRs['Type'] == 'Do not Sell or Share') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))
        filtroProcessoRestrito = (leituraDSRs['Type'] == 'Restrict Processing') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))
        filtroRevTomadaDecisao = (leituraDSRs['Type'] == 'Restrict Automated Individual Decision-making') & (~leituraDSRs['Subject Name'].str.contains('teste', case=False, na=False))

    linhas_filtradas_de_Acesso = leituraDSRs[filtroAcesso]
    linhas_filtradas_de_Eliminacao = leituraDSRs[filtroEliminacao]
    linhas_filtradas_de_DivSubProc = leituraDSRs[filtroDivSubProc]
    linhas_filtradas_de_Objecao = leituraDSRs[filtroObjecao]
    linhas_filtradas_de_Portabilidade = leituraDSRs[filtroPortabilidade]
    linhas_filtradas_de_Retificacao = leituraDSRs[filtroRetificacao]
    linhas_filtradas_de_Confirmacao = leituraDSRs[filtroConfirmacao]
    linhas_filtradas_de_NaoVenda = leituraDSRs[filtroNaoVenda]
    linhas_filtradas_de_ProcessoRestrito = leituraDSRs[filtroProcessoRestrito]
    linhas_filtradas_de_RevTomadaDecisao = leituraDSRs[filtroRevTomadaDecisao]

    # Contar o número de linhas resultantes
    numero_linhas_de_Acesso = len(linhas_filtradas_de_Acesso)
    numero_linhas_de_Eliminacao = len(linhas_filtradas_de_Eliminacao)
    numero_linhas_de_DivSubProc = len(linhas_filtradas_de_DivSubProc)
    numero_linhas_de_Objecao = len(linhas_filtradas_de_Objecao)
    numero_linhas_de_Portabilidade = len(linhas_filtradas_de_Portabilidade)
    numero_linhas_de_Retificacao = len(linhas_filtradas_de_Retificacao)
    numero_linhas_de_Confirmacao = len(linhas_filtradas_de_Confirmacao)
    numero_linhas_de_NaoVenda = len(linhas_filtradas_de_NaoVenda)
    numero_linhas_de_ProcessoRestrito = len(linhas_filtradas_de_ProcessoRestrito)
    numero_linhas_de_RevTomadaDecisao = len(linhas_filtradas_de_RevTomadaDecisao)
    
    # Total de linhas
    totalDSRs = (numero_linhas_de_Acesso + numero_linhas_de_Eliminacao + numero_linhas_de_DivSubProc + 
                numero_linhas_de_Objecao + numero_linhas_de_Portabilidade + numero_linhas_de_Retificacao + 
                numero_linhas_de_Confirmacao + numero_linhas_de_NaoVenda + numero_linhas_de_ProcessoRestrito)

    # Escreve o número de linhas nas células se o valor for maior que 0
    if numero_linhas_de_Eliminacao > 0:
        ws[f'B{linhainicialDSRs}'] = 'Eliminação'
        ws[f'C{linhainicialDSRs}'] = numero_linhas_de_Eliminacao
        linhainicialDSRs= linhainicialDSRs+1
    if numero_linhas_de_DivSubProc > 0:
        ws[f'B{linhainicialDSRs}'] = 'Compartilhamento'
        ws[f'C{linhainicialDSRs}'] = numero_linhas_de_DivSubProc
        linhainicialDSRs= linhainicialDSRs+1
    if numero_linhas_de_Objecao > 0:
        ws[f'B{linhainicialDSRs}'] = 'Objeção'
        ws[f'C{linhainicialDSRs}'] = numero_linhas_de_Objecao
        linhainicialDSRs= linhainicialDSRs+1
    if numero_linhas_de_Acesso > 0:
        ws[f'B{linhainicialDSRs}'] = 'Acesso'
        ws[f'C{linhainicialDSRs}'] = numero_linhas_de_Acesso
        linhainicialDSRs= linhainicialDSRs+1
    if numero_linhas_de_Portabilidade > 0:
        ws[f'B{linhainicialDSRs}'] = 'Portabilidade'
        ws[f'C{linhainicialDSRs}'] = numero_linhas_de_Portabilidade
        linhainicialDSRs= linhainicialDSRs+1
    if numero_linhas_de_Retificacao > 0:
        ws[f'B{linhainicialDSRs}'] = 'Retificação'
        ws[f'C{linhainicialDSRs}'] = numero_linhas_de_Retificacao
        linhainicialDSRs= linhainicialDSRs+1
    if numero_linhas_de_Confirmacao > 0:
        ws[f'B{linhainicialDSRs}'] = 'Confirmação'
        ws[f'C{linhainicialDSRs}'] = numero_linhas_de_Confirmacao
        linhainicialDSRs= linhainicialDSRs+1
    if numero_linhas_de_ProcessoRestrito > 0:
        ws[f'B{linhainicialDSRs}'] = "Processo Restrito"
        ws[f'C{linhainicialDSRs}'] = numero_linhas_de_ProcessoRestrito
        linhainicialDSRs= linhainicialDSRs+1
    if numero_linhas_de_RevTomadaDecisao > 0:
        ws[f'B{linhainicialDSRs}'] = "Revisão da tomada de Decisão"
        ws[f'C{linhainicialDSRs}'] = numero_linhas_de_RevTomadaDecisao
        linhainicialDSRs= linhainicialDSRs+1
    
    limpar_tela()
    if opcao == "cliente1" or opcao == "1":
        if numero_linhas_de_NaoVenda > 0:
            ws[f'B{linhainicialDSRs}'] = "Não Venda"
            ws[f'C{linhainicialDSRs}'] = numero_linhas_de_NaoVenda
        print("Total de DSRs filtradas no mês:", totalDSRs, "\n-----------------------------------")
        # Imprimir o resultado das DSRs
        print("Acesso :", numero_linhas_de_Acesso)
        print("Compartilhamento :", numero_linhas_de_DivSubProc)
        print("Confirmação :", numero_linhas_de_Confirmacao)
        print("Eliminação :", numero_linhas_de_Eliminacao)
        print("Objeção :", numero_linhas_de_Objecao)
        print("Portabilidade :", numero_linhas_de_Portabilidade)
        print("Processo Restrito :", numero_linhas_de_ProcessoRestrito)
        print("Retificação :", numero_linhas_de_Retificacao)
        print("Não Venda :", numero_linhas_de_NaoVenda) 
        ws[celulaTotal] = totalDSRs
    if opcao == "cliente2" or opcao == "2":
        totalDSRs= totalDSRs-numero_linhas_de_NaoVenda
        # Imprimir o resultado das DSRs
        print("Total de DSRs filtradas no mês:", totalDSRs, "\n-----------------------------------")
        print("Acesso :", numero_linhas_de_Acesso)
        print("Confirmação :", numero_linhas_de_Confirmacao)
        print("Compartilhamento :", numero_linhas_de_DivSubProc)
        print("Eliminação :", numero_linhas_de_Eliminacao)
        print("Revisão de Tomada de Decisão :", numero_linhas_de_RevTomadaDecisao)
        print("Não Venda :", numero_linhas_de_NaoVenda)  
        ws[celulaTotal] = totalDSRs
        ws[celulaNaoVenda] = numero_linhas_de_NaoVenda
    salvarExcel(graficoCliente,wb)
    time.sleep(3)
# Função de Filtragem Histórico de Cookies
def HistoricoCookies(leituraCookies,graficoCliente,opcao,celulaStatusConcedidos,celulaStatusIgnorados,
                     celulaStatusRetirados,celulaStatusRecusados,celulaConcedidos,celulaRetirados,
                     celulaIgnorados,celulaRecusados):
    
    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    numero_de_linhas_Concedidas = 0; numero_de_linhas_Recusadas = 0; 
    numero_de_linhas_Retiradas = 0; numero_de_linhas_Ignoradas = 0
    
    for chunk in leituraCookies:
        filtroConcedido = chunk[chunk['Consent Status'] == 'GRANTED']
        filtroRecusado = chunk [chunk['Consent Status'] == 'DECLINED']
        filtroRetirado = chunk [chunk['Consent Status'] == 'WITHDRAWN']
        filtroIgnorado = chunk [chunk['Consent Status'] == 'NOACTION']
        numero_de_linhas_Recusadas += len(filtroRecusado)
        numero_de_linhas_Concedidas += len(filtroConcedido)
        numero_de_linhas_Retiradas += len(filtroRetirado)
        numero_de_linhas_Ignoradas += len(filtroIgnorado)

    # Imprimir o número total de linhas concedidas
    limpar_tela()
    print("------------------\nCookies Filtrados\n------------------")
    print("Concedidos: ",numero_de_linhas_Concedidas)
    print("Recusados: ",numero_de_linhas_Recusadas)
    print("Retirados: ",numero_de_linhas_Retiradas)
    print("Ignorados: ",numero_de_linhas_Ignoradas)
    if opcao=="cliente3" or opcao==3:
        ws[celulaStatusConcedidos] = numero_de_linhas_Concedidas
        ws[celulaStatusRetirados] = numero_de_linhas_Retiradas
        ws[celulaStatusRecusados] = numero_de_linhas_Recusadas

        ws[celulaConcedidos] = numero_de_linhas_Concedidas
        ws[celulaRetirados] = numero_de_linhas_Retiradas
        ws[celulaRecusados] = numero_de_linhas_Recusadas
        
    else:    
        ws[celulaStatusConcedidos] = numero_de_linhas_Concedidas
        ws[celulaStatusIgnorados] = numero_de_linhas_Ignoradas
        ws[celulaStatusRetirados] = numero_de_linhas_Retiradas
        ws[celulaStatusRecusados] = numero_de_linhas_Recusadas

        ws[celulaConcedidos] = numero_de_linhas_Concedidas
        ws[celulaRetirados] = numero_de_linhas_Retiradas
        ws[celulaIgnorados] = numero_de_linhas_Ignoradas
        ws[celulaRecusados] = numero_de_linhas_Recusadas
    
    salvarExcel(graficoCliente,wb)
    time.sleep(3)
# Função de Filtragem Histórico de Cookies Cliente2 Segundo Ambiente
def HistoricoCookiesCliente2Ambiente(leituraCookiesCliente2Ambiente,graficoCliente):
    
    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    
    numero_de_linhas_Concedidas = 0; numero_de_linhas_Recusadas = 0; 
    numero_de_linhas_Retiradas = 0; numero_de_linhas_Ignoradas = 0

    for chunk in leituraCookiesCliente2Ambiente:
        filtroConcedido = chunk[chunk['Consent Status'] == 'GRANTED']
        filtroRecusado = chunk [chunk['Consent Status'] == 'DECLINED']
        filtroRetirado = chunk [chunk['Consent Status'] == 'WITHDRAWN']
        filtroIgnorado = chunk [chunk['Consent Status'] == 'NOACTION']
        numero_de_linhas_Recusadas += len(filtroRecusado)
        numero_de_linhas_Concedidas += len(filtroConcedido)
        numero_de_linhas_Retiradas += len(filtroRetirado)
        numero_de_linhas_Ignoradas += len(filtroIgnorado)

    # Imprimir o número total de linhas concedidas
    limpar_tela()
    print("-----------------------\nCookies Filtrados (Cliente2 Segundo Ambiente)\n-----------------------")
    print("Concedidos: ",numero_de_linhas_Concedidas)
    print("Recusados: ",numero_de_linhas_Recusadas)
    print("Retirados: ",numero_de_linhas_Retiradas)
    print("Ignorados: ",numero_de_linhas_Ignoradas)

    ws[celulaStatusConcedidosCliente2Ambiente] = numero_de_linhas_Concedidas
    ws[celulaStatusIgnoradosCliente2Ambiente] = numero_de_linhas_Ignoradas
    ws[celulaStatusRetiradosCliente2Ambiente] = numero_de_linhas_Retiradas
    ws[celulaStatusRecusadosCliente2Ambiente] = numero_de_linhas_Recusadas

    salvarExcel(graficoCliente,wb)
    time.sleep(3)
# Função de Filtragem das Origem dos Consentimento de Cookies Cliente 1
def OrigemConsentimentoCookiesCliente1(leituraCookies, graficoCliente,origensCliente): 
    # Carregar arquivo Excel
    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()

    # Processar cookies por chunk
    for chunk in leituraCookies:
        for origem, urls in agrupamentos.items():
            # Somar as ocorrências de todas as URLs dentro de um grupo
            total_linhas = 0
            for url in urls:
                # Para cada URL no agrupamento, contar quantas vezes ela aparece no chunk
                total_linhas += len(chunk[chunk['Location Code'] == url])
            # Atualizar o dicionário com o total de linhas para essa origem
            origensCliente[origem] += total_linhas
    
    # Ordenar as origens por número de ocorrências
    origensClienteOrdenadas = sorted(origensCliente.items(), key=lambda x: x[1], reverse=True)

    # Imprimir e salvar resultados
    print("\n-----------------------------------\n Origem Dos Consentimentos (Sites) \n-----------------------------------")
    
    # Alteração aqui: usar len(origensClienteOrdenadas) para determinar o número de linhas a exibir
    for i, (origem, linhas) in enumerate(origensClienteOrdenadas[:len(origensClienteOrdenadas)], start=linhaInicialSitesCliente1):
        if linhas > 0:
            ws[f'B{i}'] = origem
            ws[f'C{i}'] = linhas
            print(f"{origem}: {linhas}")
        
    salvarExcel(graficoCliente, wb)
    time.sleep(3)  # Pausa para visualização
# Função de Filtragem das Origem dos Consentimento de Cookies Cliente2
def OrigemConsentimentoCookiesCliente2(leituraCookies, leituraCookiesCliente2Ambiente, origensCliente, graficoCliente): 
    # Carregar arquivo Excel
    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    # Processar os cookies do primeiro arquivo (leituraCookies)
    for chunk in leituraCookies:
        for origem in origensCliente.keys():  # Itera sobre todas as origens
            filtro = chunk[chunk['Location Code'] == origem]  # Filtra os dados por origem
            origensCliente[origem] += len(filtro)  # Soma as ocorrências para cada origem

    # Processar os cookies do segundo arquivo (leituraCookiesCliente2Ambiente)
    for chunk in leituraCookiesCliente2Ambiente:
        filtroCliente2Ambiente = chunk[chunk['Location Code'] == 'https://Cliente2Ambiente.com.br/']
        origensCliente['https://Cliente2Ambiente.com.br/'] += len(filtroCliente2Ambiente)

    # Ordena as origens por número de ocorrências
    origensClienteOrdenadas = sorted(origensCliente.items(), key=lambda x: x[1], reverse=True)
    limpar_tela()
    print("-----------------------------------\n Origem Dos Consentimentos (Sites) \n-----------------------------------")

    # Gravar os resultados no Excel e imprimir na tela
    for i, (nome, linhas) in enumerate(origensClienteOrdenadas[:8], start=linhaInicialSitesCliente2):
        if linhas > 0:
            ws[f'B{i}'] = nome
            ws[f'C{i}'] = linhas
            print(f"{nome}: {linhas}")

    # Salvar as alterações no arquivo Excel
    salvarExcel(graficoCliente, wb)
    time.sleep(3)  # Pausa para visualização
# Função de Filtragem das Origem dos Consentimento de Cookies Cliente3
def OrigemConsentimentoCookiesCliente3(leituraCookies, graficoCliente,origensCliente): 
    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    # Processar os cookies
    for chunk in leituraCookies:
        for origem in origensCliente.keys():  # Itera sobre todas as origens no dicionário
            filtro = chunk[chunk['Location Code'] == origem]  # Filtra os dados por origem
            origensCliente[origem] += len(filtro)  # Soma as ocorrências para cada origem

    # Ordenar as origens por número de ocorrências
    origensClienteOrdenadas = sorted(origensCliente.items(), key=lambda x: x[1], reverse=True)
    limpar_tela()
    print("-----------------------------------\n Origem Dos Consentimentos (Sites) \n-----------------------------------")

    # Gravar os resultados no Excel e imprimir na tela
    for i, (nome, linhas) in enumerate(origensClienteOrdenadas[:2], start=linhaInicialSitesCliente3):
        if linhas > 0:
            ws[f'B{i}'] = nome
            ws[f'C{i}'] = linhas
            print(nome, ": ", linhas)

    # Salvar as alterações no arquivo Excel
    salvarExcel(graficoCliente, wb)
    time.sleep(3)  # Pausa para visualização
# Função de Classificação de Cookies Cliente1
def ClassificacaoCookies(leituraCookies, graficoCliente, ConcedidoEssencial, RecusadoEssencial, 
                            RetiradoEssencial, ConcedidoPublicidade, RecusadoPublicidade, 
                            RetiradoPublicidade, ConcedidoAnaliseEPersonalizacao, 
                            RecusadoAnaliseEPersonalizacao, RetiradoAnaliseEPersonalizacao, 
                            ConcedidoDesempenhoEFuncionalidade, RecusadoDesempenhoEFuncionalidade, 
                            RetiradoDesempenhoEFuncionalidade):

    numero_de_filtroConcedidoEssencial=0; numero_de_filtroConcedidoPublicidade=0; numero_de_filtroRecusadoPublicidade=0; numero_de_filtroRetiradoPublicidade=0
    numero_de_filtroConcedidoAnaliseEPersonalizacao=0; numero_de_filtroRecusadoAnaliseEPersonalizacao=0; numero_de_filtroRetiradoAnaliseEPersonalizacao=0
    numero_de_filtroConcedidoDesempenhoEFuncionalidade=0; numero_de_filtroRecusadoDesempenhoEFuncionalidade=0; numero_de_filtroRetiradoDesempenhoEFuncionalidade=0

    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()

    for chunk in leituraCookies:
        filtroConcedidoEssencial = chunk[(chunk['Consent Status'] == 'GRANTED') & ((chunk['Cookie Category'] == "Essencial") | (chunk['Cookie Category'] == "Essential"))]

        filtroConcedidoPublicidade = chunk[(chunk['Consent Status'] == 'GRANTED') & ((chunk['Cookie Category'] == "Publicidade") | (chunk['Cookie Category'] == "Advertising"))]
        filtroRecusadoPublicidade = chunk[(chunk['Consent Status'] == 'DECLINED') & ((chunk['Cookie Category'] == "Publicidade") | (chunk['Cookie Category'] == "Advertising"))]
        filtroRetiradoPublicidade = chunk[(chunk['Consent Status'] == 'WITHDRAWN') & ((chunk['Cookie Category'] == "Publicidade") | (chunk['Cookie Category'] == "Advertising"))]

        filtroConcedidoAnaliseEPersonalizacao = chunk[(chunk['Consent Status'] == 'GRANTED') & ((chunk['Cookie Category'] == "AnÃ¡lise e personalizaÃ§Ã£o") | (chunk['Cookie Category'] == "Analytics & Customization"))]
        filtroRecusadoAnaliseEPersonalizacao = chunk[(chunk['Consent Status'] == 'DECLINED') & ((chunk['Cookie Category'] == "AnÃ¡lise e personalizaÃ§Ã£o") | (chunk['Cookie Category'] == "Analytics & Customization"))]
        filtroRetiradoAnaliseEPersonalizacao = chunk[(chunk['Consent Status'] == 'WITHDRAWN') & ((chunk['Cookie Category'] == "AnÃ¡lise e personalizaÃ§Ã£o") | (chunk['Cookie Category'] == "Analytics & Customization"))]

        filtroConcedidoDesempenhoEFuncionalidade = chunk[(chunk['Consent Status'] == 'GRANTED') & ((chunk['Cookie Category'] == "Desempenho e funcionalidade") | (chunk['Cookie Category'] == "Performance & Functionality"))]
        filtroRecusadoDesempenhoEFuncionalidade = chunk[(chunk['Consent Status'] == 'DECLINED') & ((chunk['Cookie Category'] == "Desempenho e funcionalidade") | (chunk['Cookie Category'] == "Performance & Functionality"))]
        filtroRetiradoDesempenhoEFuncionalidade = chunk[(chunk['Consent Status'] == 'WITHDRAWN') & ((chunk['Cookie Category'] == "Desempenho e funcionalidade") | (chunk['Cookie Category'] == "Performance & Functionality"))]
        
        numero_de_filtroConcedidoEssencial += len(filtroConcedidoEssencial)
        numero_de_filtroConcedidoPublicidade += len(filtroConcedidoPublicidade)
        numero_de_filtroRecusadoPublicidade += len(filtroRecusadoPublicidade)
        numero_de_filtroRetiradoPublicidade += len(filtroRetiradoPublicidade)
        numero_de_filtroConcedidoAnaliseEPersonalizacao += len(filtroConcedidoAnaliseEPersonalizacao)
        numero_de_filtroRecusadoAnaliseEPersonalizacao += len(filtroRecusadoAnaliseEPersonalizacao)
        numero_de_filtroRetiradoAnaliseEPersonalizacao += len(filtroRetiradoAnaliseEPersonalizacao)
        numero_de_filtroConcedidoDesempenhoEFuncionalidade += len(filtroConcedidoDesempenhoEFuncionalidade)
        numero_de_filtroRecusadoDesempenhoEFuncionalidade += len(filtroRecusadoDesempenhoEFuncionalidade)
        numero_de_filtroRetiradoDesempenhoEFuncionalidade += len(filtroRetiradoDesempenhoEFuncionalidade)

    limpar_tela()
    print("---------------- Cookies Classificados (Classificados, Recusados e Retirados) ----------------")
    print("Essencial: ", numero_de_filtroConcedidoEssencial," | 0 | 0")
    print("Publicidade: ", numero_de_filtroConcedidoPublicidade," | ", numero_de_filtroRecusadoPublicidade, " | ", numero_de_filtroRetiradoPublicidade) 
    print("Análise e Personalização: ", numero_de_filtroConcedidoAnaliseEPersonalizacao, " | ", numero_de_filtroRecusadoAnaliseEPersonalizacao, " | ", numero_de_filtroRetiradoAnaliseEPersonalizacao)
    print("Desempenho e Funcionalidade: ", numero_de_filtroConcedidoDesempenhoEFuncionalidade, " | ", numero_de_filtroRecusadoDesempenhoEFuncionalidade, " | ", numero_de_filtroRetiradoDesempenhoEFuncionalidade)
    
    ws[ConcedidoEssencial] = numero_de_filtroConcedidoEssencial 
    ws[RecusadoEssencial] = "0"
    ws[RetiradoEssencial] = "0"

    ws[ConcedidoPublicidade] = numero_de_filtroConcedidoPublicidade
    ws[RecusadoPublicidade] = numero_de_filtroRecusadoPublicidade
    ws[RetiradoPublicidade] = numero_de_filtroRetiradoPublicidade

    ws[ConcedidoAnaliseEPersonalizacao] = numero_de_filtroConcedidoAnaliseEPersonalizacao
    ws[RecusadoAnaliseEPersonalizacao] = numero_de_filtroRecusadoAnaliseEPersonalizacao
    ws[RetiradoAnaliseEPersonalizacao] = numero_de_filtroRetiradoAnaliseEPersonalizacao

    ws[ConcedidoDesempenhoEFuncionalidade] = numero_de_filtroConcedidoDesempenhoEFuncionalidade
    ws[RecusadoDesempenhoEFuncionalidade] = numero_de_filtroRecusadoDesempenhoEFuncionalidade
    ws[RetiradoDesempenhoEFuncionalidade] = numero_de_filtroRetiradoDesempenhoEFuncionalidade

    # Salvar as alterações no arquivo Excel
    salvarExcel(graficoCliente, wb)
    time.sleep(3)
# Função de Filtragem dos Sistemas Operacionais
def SistemasOperacionaisCookies(leituraCookies, linhainicial, graficoCliente): 
    numero_de_linhas_Android = 0; numero_de_linhas_ArchLinux = 0; numero_de_linhas_AtvOSX = 0; numero_de_linhas_Bada = 0
    numero_de_linhas_BlackberryOs = 0; numero_de_linhas_BlackberryTabletOs = 0; numero_de_linhas_Brew = 0; numero_de_linhas_BrewMP = 0
    numero_de_linhas_Bsd = 0; numero_de_linhas_Centos = 0; numero_de_linhas_ChromeOS = 0; numero_de_linhas_ChromeCast = 0
    numero_de_linhas_Daiko = 0; numero_de_linhas_Debian = 0; numero_de_linhas_Fedora = 0; numero_de_linhas_FirefoxOs = 0
    numero_de_linhas_FreeBSD = 0; numero_de_linhas_Fu = 0; numero_de_linhas_Gentoo = 0; numero_de_linhas_GoogleTv = 0; numero_de_linhas_Hisense = 0
    numero_de_linhas_Ios = 0; numero_de_linhas_Jvc = 0; numero_de_linhas_Kaios = 0; numero_de_linhas_Kindle = 0; numero_de_linhas_Kubuntu = 0
    numero_de_linhas_Linux = 0; numero_de_linhas_LinuxMint = 0; numero_de_linhas_MacOs = 0
    numero_de_linhas_MacOsX = 0; numero_de_linhas_Maemo = 0; numero_de_linhas_Mageia = 0; numero_de_linhas_Mandriva = 0
    numero_de_linhas_Meego = 0; numero_de_linhas_NetBsd = 0; numero_de_linhas_NokiaSeries40 = 0; numero_de_linhas_OpenBsd = 0
    numero_de_linhas_OpenSuse = 0; numero_de_linhas_Other = 0; numero_de_linhas_Panasonic = 0; numero_de_linhas_Philips = 0; numero_de_linhas_Qilive = 0
    numero_de_linhas_RedHat = 0; numero_de_linhas_Regal = 0; numero_de_linhas_Roku = 0; numero_de_linhas_Sailfish = 0; numero_de_linhas_Slackware = 0
    numero_de_linhas_Solaris = 0; numero_de_linhas_Sony = 0; numero_de_linhas_Suse = 0; numero_de_linhas_SymbianOs = 0; numero_de_linhas_Symbian3 = 0
    numero_de_linhas_Symbian3Anna = 0; numero_de_linhas_Symbian3Belle = 0; numero_de_linhas_Tizen = 0; numero_de_linhas_Ubuntu = 0
    numero_de_linhas_Vre = 0; numero_de_linhas_Web0s = 0; numero_de_linhas_Webos = 0; numero_de_linhas_Webtv = 0
    numero_de_linhas_Windows = 0; numero_de_linhas_WindowsMobile = 0; numero_de_linhas_WindowsPhone = 0

    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    
    for chunk in leituraCookies:
        
        filtroAndroid = chunk[chunk['Client Os Family'] == 'android']
        filtroArchLinux = chunk[chunk['Client Os Family'] == 'arch linux']
        filtroAtvOsX = chunk[chunk['Client Os Family'] == 'atv os x']
        filtroBada = chunk[chunk['Client Os Family'] == 'bada']
        filtroBlackberryOs = chunk[chunk['Client Os Family'] == 'blackberry os']
        filtroBlackberryTabletOs = chunk[chunk['Client Os Family'] == 'blackberry tablet os']
        filtroBrew = chunk[chunk['Client Os Family'] == 'brew']
        filtroBrewMp = chunk[chunk['Client Os Family'] == 'brew mp']
        filtroBsd = chunk[chunk['Client Os Family'] == 'bsd']
        filtroCentos = chunk[chunk['Client Os Family'] == 'centos']
        filtroChromeOs = chunk[chunk['Client Os Family'] == 'chrome os']
        filtroChromeCast = chunk[chunk['Client Os Family'] == 'chromecast']
        filtroDaiko = chunk[chunk['Client Os Family'] == 'daiko']
        filtroDebian = chunk[chunk['Client Os Family'] == 'debian']
        filtroFedora = chunk[chunk['Client Os Family'] == 'fedora']
        filtroFirefoxOs = chunk[chunk['Client Os Family'] == 'firefox os']
        filtroFreeBSD = chunk[chunk['Client Os Family'] == 'freebsd']
        filtroFu = chunk[chunk['Client Os Family'] == 'fu']
        filtroGentoo = chunk[chunk['Client Os Family'] == 'gentoo']
        filtroGoogleTv = chunk[chunk['Client Os Family'] == 'googletv']
        filtroHisense = chunk[chunk['Client Os Family'] == 'hisense']
        filtroIos = chunk[chunk['Client Os Family'] == 'ios']
        filtroJvc = chunk[chunk['Client Os Family'] == 'jvc']
        filtroKaios = chunk[chunk['Client Os Family'] == 'kaios']
        filtroKindle = chunk[chunk['Client Os Family'] == 'kindle']
        filtroKubuntu = chunk[chunk['Client Os Family'] == 'kubuntu']
        filtroLinux = chunk[chunk['Client Os Family'] == 'linux']
        filtroLinuxMint = chunk[chunk['Client Os Family'] == 'linux mint']
        filtroMacOs = chunk[chunk['Client Os Family'] == 'mac os']
        filtroMacOsX = chunk[chunk['Client Os Family'] == 'mac os x']
        filtroMaemo = chunk[chunk['Client Os Family'] == 'maemo']
        filtroMageia = chunk[chunk['Client Os Family'] == 'mageia']
        filtroMandriva = chunk[chunk['Client Os Family'] == 'mandriva']
        filtroMeego = chunk[chunk['Client Os Family'] == 'meego']
        filtroNetBsd = chunk[chunk['Client Os Family'] == 'netbsd']
        filtroNokiaSeries40 = chunk[chunk['Client Os Family'] == 'nokia series 40']
        filtroOpenBsd = chunk[chunk['Client Os Family'] == 'openbsd']
        filtroOpenSuse = chunk[chunk['Client Os Family'] == 'opensuse']
        filtroOther = chunk[chunk['Client Os Family'] == 'other']
        filtroPanasonic = chunk[chunk['Client Os Family'] == 'panasonic']
        filtroPhilips = chunk[chunk['Client Os Family'] == 'philips']
        filtroQilive = chunk[chunk['Client Os Family'] == 'qilive']
        filtroRedHat = chunk[chunk['Client Os Family'] == 'red hat']
        filtroRegal = chunk[chunk['Client Os Family'] == 'regal']
        filtroRoku = chunk[chunk['Client Os Family'] == 'roku']
        filtroSailfish = chunk[chunk['Client Os Family'] == 'sailfish']
        filtroSlackware = chunk[chunk['Client Os Family'] == 'slackware']
        filtroSolaris = chunk[chunk['Client Os Family'] == 'solaris']
        filtroSony = chunk[chunk['Client Os Family'] == 'sony']
        filtroSuse = chunk[chunk['Client Os Family'] == 'suse']
        filtroSymbianOs = chunk[chunk['Client Os Family'] == 'symbian os']
        filtroSymbian3 = chunk[chunk['Client Os Family'] == 'symbian^3']
        filtroSymbian3Anna = chunk[chunk['Client Os Family'] == 'symbian^3 anna']
        filtroSymbian3Belle = chunk[chunk['Client Os Family'] == 'symbian^3 belle']
        filtroTizen = chunk[chunk['Client Os Family'] == 'tizen']
        filtroUbuntu = chunk[chunk['Client Os Family'] == 'ubuntu']
        filtroVre = chunk[chunk['Client Os Family'] == 'vre']
        filtroWeb0s = chunk[chunk['Client Os Family'] == 'web0s']
        filtroWebos = chunk[chunk['Client Os Family'] == 'webos']
        filtroWebTv = chunk[chunk['Client Os Family'] == 'webtv']
        filtroWindows = chunk[chunk['Client Os Family'] == 'windows']
        filtroWindowsMobile = chunk[chunk['Client Os Family'] == 'windows mobile']
        filtroWindowsPhone = chunk[chunk['Client Os Family'] == 'windows phone']
        
        numero_de_linhas_Android += len(filtroAndroid)
        numero_de_linhas_ArchLinux += len(filtroArchLinux)
        numero_de_linhas_AtvOSX += len(filtroAtvOsX)
        numero_de_linhas_Bada += len(filtroBada)
        numero_de_linhas_BlackberryOs += len(filtroBlackberryOs)
        numero_de_linhas_BlackberryTabletOs += len(filtroBlackberryTabletOs)
        numero_de_linhas_Brew += len(filtroBrew)
        numero_de_linhas_BrewMP += len(filtroBrewMp)
        numero_de_linhas_Bsd += len(filtroBsd)
        numero_de_linhas_Centos += len(filtroCentos)
        numero_de_linhas_ChromeOS += len(filtroChromeOs)
        numero_de_linhas_ChromeCast += len(filtroChromeCast)
        numero_de_linhas_Daiko += len(filtroDaiko)
        numero_de_linhas_Debian += len(filtroDebian)
        numero_de_linhas_Fedora += len(filtroFedora)
        numero_de_linhas_FirefoxOs += len(filtroFirefoxOs)
        numero_de_linhas_FreeBSD += len(filtroFreeBSD)
        numero_de_linhas_Fu += len(filtroFu)
        numero_de_linhas_Gentoo += len(filtroGentoo)
        numero_de_linhas_GoogleTv += len(filtroGoogleTv)
        numero_de_linhas_Hisense += len(filtroHisense)
        numero_de_linhas_Ios += len(filtroIos)
        numero_de_linhas_Jvc += len(filtroJvc)
        numero_de_linhas_Kaios += len(filtroKaios)
        numero_de_linhas_Kindle += len(filtroKindle)
        numero_de_linhas_Kubuntu += len(filtroKubuntu)
        numero_de_linhas_Linux += len(filtroLinux)
        numero_de_linhas_LinuxMint += len(filtroLinuxMint)
        numero_de_linhas_MacOs += len(filtroMacOs)
        numero_de_linhas_MacOsX += len(filtroMacOsX)
        numero_de_linhas_Maemo += len(filtroMaemo)
        numero_de_linhas_Mageia += len(filtroMageia)
        numero_de_linhas_Mandriva += len(filtroMandriva)
        numero_de_linhas_Meego += len(filtroMeego)
        numero_de_linhas_NetBsd += len(filtroNetBsd)
        numero_de_linhas_NokiaSeries40 += len(filtroNokiaSeries40)
        numero_de_linhas_OpenBsd += len(filtroOpenBsd)
        numero_de_linhas_OpenSuse += len(filtroOpenSuse)
        numero_de_linhas_Other += len(filtroOther)
        numero_de_linhas_Panasonic += len(filtroPanasonic)
        numero_de_linhas_Philips += len(filtroPhilips)
        numero_de_linhas_Qilive += len(filtroQilive)
        numero_de_linhas_RedHat += len(filtroRedHat)
        numero_de_linhas_Regal += len(filtroRegal)
        numero_de_linhas_Roku += len(filtroRoku)
        numero_de_linhas_Sailfish += len(filtroSailfish)
        numero_de_linhas_Slackware += len(filtroSlackware)
        numero_de_linhas_Solaris += len(filtroSolaris)
        numero_de_linhas_Sony += len(filtroSony)
        numero_de_linhas_Suse += len(filtroSuse)
        numero_de_linhas_SymbianOs += len(filtroSymbianOs)
        numero_de_linhas_Symbian3 += len(filtroSymbian3)
        numero_de_linhas_Symbian3Anna += len(filtroSymbian3Anna)
        numero_de_linhas_Symbian3Belle += len(filtroSymbian3Belle)
        numero_de_linhas_Tizen += len(filtroTizen)
        numero_de_linhas_Ubuntu += len(filtroUbuntu)
        numero_de_linhas_Vre += len(filtroVre)
        numero_de_linhas_Web0s += len(filtroWeb0s)
        numero_de_linhas_Webos += len(filtroWebos)
        numero_de_linhas_Webtv += len(filtroWebTv)
        numero_de_linhas_Windows += len(filtroWindows)
        numero_de_linhas_WindowsMobile += len(filtroWindowsMobile)
        numero_de_linhas_WindowsPhone += len(filtroWindowsPhone)
                
    linhas_por_sistema = {
    "Android": numero_de_linhas_Android,
	"ArchLinux": numero_de_linhas_ArchLinux,
    "Atv OS X": numero_de_linhas_AtvOSX,
	"Bada": numero_de_linhas_Bada,
	"Blackberry OS": numero_de_linhas_BlackberryOs,
	"Blackberry Tablet OS": numero_de_linhas_BlackberryTabletOs,
	"Brew": numero_de_linhas_Brew,
	"Brew MP": numero_de_linhas_BrewMP,
	"BSD": numero_de_linhas_Bsd,
	"Centos": numero_de_linhas_Centos,
	"Chrome OS": numero_de_linhas_ChromeOS,
	"Chrome Cast": numero_de_linhas_ChromeCast,
	"Daiko": numero_de_linhas_Daiko,
	"Debian": numero_de_linhas_Debian,
	"Fedora": numero_de_linhas_Fedora,
	"Firefox OS": numero_de_linhas_FirefoxOs,
	"Free BSD": numero_de_linhas_FreeBSD,
	"Fu": numero_de_linhas_Fu,
	"Gentoo": numero_de_linhas_Gentoo,
	"Google TV": numero_de_linhas_GoogleTv,
	"Hisense": numero_de_linhas_Hisense,
	"IOS": numero_de_linhas_Ios,
	"JVC": numero_de_linhas_Jvc,
	"Kaios": numero_de_linhas_Kaios,
	"Kindle": numero_de_linhas_Kindle,
	"Kubuntu": numero_de_linhas_Kubuntu,
	"Linux": numero_de_linhas_Linux,
	"Linux Mint": numero_de_linhas_LinuxMint,
	"Mac OS": numero_de_linhas_MacOs,
	"Mac OS X": numero_de_linhas_MacOsX,
	"Maemo": numero_de_linhas_Maemo,
	"Mageia": numero_de_linhas_Mageia,
	"Mandriva": numero_de_linhas_Mandriva,
	"Meego": numero_de_linhas_Meego,
	"Net BSD": numero_de_linhas_NetBsd, 
	"Nokia Series 40": numero_de_linhas_NokiaSeries40,
	"Open BSD": numero_de_linhas_OpenBsd,
	"Open Suse": numero_de_linhas_OpenSuse,
	"Other": numero_de_linhas_Other,
	"Panasonic": numero_de_linhas_Panasonic,
	"Philips": numero_de_linhas_Philips,
	"Qilive": numero_de_linhas_Qilive,
	"Red Hat": numero_de_linhas_RedHat,
	"Regal": numero_de_linhas_Regal,
	"Roku": numero_de_linhas_Roku,
	"Sailfish": numero_de_linhas_Sailfish,
	"Slackware": numero_de_linhas_Slackware,
	"Solaris": numero_de_linhas_Solaris,
	"Sony": numero_de_linhas_Sony,
	"Suse": numero_de_linhas_Suse,
	"Symbian OS": numero_de_linhas_SymbianOs,
	"Symbian^3": numero_de_linhas_Symbian3,
	"Symbian^3 anna": numero_de_linhas_Symbian3Anna,
	"Symbian^3 belle": numero_de_linhas_Symbian3Belle,
	"Tizen": numero_de_linhas_Tizen,
	"Ubuntu": numero_de_linhas_Ubuntu,
	"VRE": numero_de_linhas_Vre,
	"Web0s": numero_de_linhas_Web0s,
	"Webos": numero_de_linhas_Webos,
	"Web TV": numero_de_linhas_Webtv,
	"Windows": numero_de_linhas_Windows,
	"Windows Mobile": numero_de_linhas_WindowsMobile,
	"Windows Phone": numero_de_linhas_WindowsPhone
    }
    # Ordenar o dicionário por número de linhas em ordem decrescente
    linhas_por_sistema_ordenadas = sorted(linhas_por_sistema.items(), key=lambda x: x[1], reverse=True)
    limpar_tela()
    # Mostrar o dicionário ordenado
    print("-------------------------------------------\n 7 Maiores Sistemas Operacionais Filtrados\n-------------------------------------------")
    for i, (os, linhas) in enumerate(linhas_por_sistema_ordenadas[:7], start=linhainicial):
        ws[f'B{i}'] = os
        ws[f'C{i}'] = linhas
        print(os,": ",linhas)
    
    # Salvar as alterações no arquivo Excel
    salvarExcel(graficoCliente, wb)
    time.sleep(3)
# Função das Origens de Consentimento Universal Cliente2
def OrigemConsentimentoUniversalCliente2(leituraConsentimentoUniversal,graficoCliente):
    numero_de_linhas_Cliente2 = 0; numero_de_linhas_PortalCliente2 = 0; 

    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    for chunk in leituraConsentimentoUniversal:
        Cliente2 = chunk[chunk['Consent Source Name'] == 'Cliente2']
        PortalCliente2 = chunk[chunk['Consent Source Name'] == 'PortalCliente2']
    

        numero_de_linhas_Cliente2 += len(Cliente2)
        numero_de_linhas_PortalCliente2 += len(PortalCliente2)

    origens = {
        "Cliente 2": numero_de_linhas_Cliente2,
        "Portal Cliente 2 ": numero_de_linhas_PortalCliente2,
    }

    linhas_por_origens = sorted(origens.items(), key=lambda x: x[1], reverse=True)

    limpar_tela()
    print ("----------------------------------------\n Origem Dos Consentimentos Universais \n----------------------------------------")
    # Imprimir o número Origens de Consentimentos
    for i, (nome, linhas) in enumerate(linhas_por_origens[:8], start=linhaInicialConsentimentoUniversalCliente3):
        ws[f'B{i}'] = nome
        ws[f'C{i}'] = linhas
        print(nome,": ",linhas)
    
    # Salvar as alterações no arquivo Excel
    salvarExcel(graficoCliente,wb)
    time.sleep(3)
#Função para Pegar o Histórico/Status de Consentimento Universal 
def HistoricoConsentimentoUniversal(leituraConsentimentoUniversal,graficoCliente,ConsentimentoConcedido,
                                          ConsentimentoRetirado,ConsentimentoRecusado):
    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    numero_de_linhas_Concedidas = 0; numero_de_linhas_Recusadas = 0; numero_de_linhas_Retiradas = 0

    for chunk in leituraConsentimentoUniversal:
        filtroConcedido = chunk[chunk['Consent Status'] == 'GRANTED']
        filtroRecusado = chunk [chunk['Consent Status'] == 'DECLINED']
        filtroRetirado = chunk [chunk['Consent Status'] == 'WITHDRAWN']

        numero_de_linhas_Recusadas += len(filtroRecusado)
        numero_de_linhas_Concedidas += len(filtroConcedido)
        numero_de_linhas_Retiradas += len(filtroRetirado)

    # Imprimir o número total de linhas concedidas
    limpar_tela()
    print("-----------------------------------------\nHistórico dos Consentimentos Universais\n-----------------------------------------")
    print("Concedidos: ",numero_de_linhas_Concedidas)
    print("Recusados: ",numero_de_linhas_Recusadas)
    print("Retirados: ",numero_de_linhas_Retiradas)

    ws[ConsentimentoConcedido] = numero_de_linhas_Concedidas
    ws[ConsentimentoRetirado] = numero_de_linhas_Retiradas
    ws[ConsentimentoRecusado] = numero_de_linhas_Recusadas

    # Salvar as alterações no arquivo Excel
    salvarExcel(graficoCliente,wb)
    time.sleep(3)
#Função para Pegar a Finalidade de Consentimento Universal Cliente 2
def FinalidadeConsentimentoUniversalCliente2 (leituraConsentimentoUniversal,graficoCliente):
    numero_de_WhatsappConcedido=0; numero_de_WhatsappRecusado=0; numero_de_WhatsappRetirado=0
    numero_de_SMSConcedido=0; numero_de_SMSRecusado=0; numero_de_SMSRetirado=0
    numero_de_EmailConcedido=0; numero_de_EmailRecusado=0; numero_de_EmailRetirado=0
    numero_de_TelefoneConcedido=0; numero_de_TelefoneRecusado=0; numero_de_TelefoneRetirado=0
    numero_de_NotificacoesConcedido=0; numero_de_NotificacoesRecusado=0; numero_de_NotificacoesRetirado=0
    numero_de_OfertasConcedido = 0; numero_de_OfertasRecusado=0; numero_de_OfertasRetirado=0

    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    for chunk in leituraConsentimentoUniversal:
        filtroTelefoneConcedido = chunk[(chunk['Consent Status'] == 'GRANTED') & (chunk['Consent Purpose'] == "Telefone")]
        filtroTelefoneRecusado = chunk[(chunk['Consent Status'] == 'DECLINED') & (chunk['Consent Purpose'] == "Telefone")]
        filtroTelefoneRetirado = chunk[(chunk['Consent Status'] == 'WITHDRAWN') & (chunk['Consent Purpose'] == "Telefone")]

        filtroSMSConcedido = chunk[(chunk['Consent Status'] == 'GRANTED') & (chunk['Consent Purpose'] == "SMS ")]
        filtroSMSRecusado = chunk[(chunk['Consent Status'] == 'DECLINED') & (chunk['Consent Purpose'] == "SMS ")]
        filtroSMSRetirado = chunk[(chunk['Consent Status'] == 'WITHDRAWN') & (chunk['Consent Purpose'] == "SMS ")]

        filtroEmailConcedido = chunk[(chunk['Consent Status'] == 'GRANTED') & (chunk['Consent Purpose'] == "E-mail")]
        filtroEmailRecusado = chunk[(chunk['Consent Status'] == 'DECLINED') & (chunk['Consent Purpose'] == "E-mail")]
        filtroEmailRetirado = chunk[(chunk['Consent Status'] == 'WITHDRAWN') & (chunk['Consent Purpose'] == "E-mail")]

        filtroWhatsappConcedido = chunk[(chunk['Consent Status'] == 'GRANTED') & (chunk['Consent Purpose'] == "Whatsapp")]
        filtroWhatsappRecusado = chunk[(chunk['Consent Status'] == 'DECLINED') & (chunk['Consent Purpose'] == "Whatsapp")]
        filtroWhatsappRetirado = chunk[(chunk['Consent Status'] == 'WITHDRAWN') & (chunk['Consent Purpose'] == "Whatsapp")]

        filtroOfertasConcedido = chunk[(chunk['Consent Status'] == 'GRANTED') & (chunk['Consent Purpose'] == "Ofertas de produtos Cliente 2")]
        filtroOfertasRecusado = chunk[(chunk['Consent Status'] == 'DECLINED') & (chunk['Consent Purpose'] == "Ofertas de produtos Cliente 2")]
        filtroOfertasRetirado = chunk[(chunk['Consent Status'] == 'WITHDRAWN') & (chunk['Consent Purpose'] == "Ofertas de produtos Cliente 2")]

        filtroNotificacoesConcedido = chunk[(chunk['Consent Status'] == 'GRANTED') & (chunk['Consent Purpose'] == "NotificaÃ§Ãµes")]
        filtroNotificacoesRecusado = chunk[(chunk['Consent Status'] == 'DECLINED') & (chunk['Consent Purpose'] == "NotificaÃ§Ãµes")]
        filtroNotificacoesRetirado = chunk[(chunk['Consent Status'] == 'WITHDRAWN') & (chunk['Consent Purpose'] == "NotificaÃ§Ãµes")]
        
        numero_de_WhatsappConcedido += len(filtroWhatsappConcedido)
        numero_de_WhatsappRecusado += len(filtroWhatsappRecusado)
        numero_de_WhatsappRetirado += len(filtroWhatsappRetirado)

        numero_de_SMSConcedido += len(filtroSMSConcedido)
        numero_de_SMSRecusado += len(filtroSMSRecusado)
        numero_de_SMSRetirado += len(filtroSMSRetirado)

        numero_de_TelefoneConcedido += len(filtroTelefoneConcedido)
        numero_de_TelefoneRecusado += len(filtroTelefoneRecusado)
        numero_de_TelefoneRetirado += len(filtroTelefoneRetirado)
        
        numero_de_EmailConcedido += len(filtroEmailConcedido)
        numero_de_EmailRecusado += len(filtroEmailRecusado)
        numero_de_EmailRetirado += len(filtroEmailRetirado)

        numero_de_NotificacoesConcedido += len(filtroNotificacoesConcedido)
        numero_de_NotificacoesRecusado += len(filtroNotificacoesRecusado)
        numero_de_NotificacoesRetirado += len(filtroNotificacoesRetirado)

        numero_de_OfertasConcedido += len(filtroOfertasConcedido)
        numero_de_OfertasRecusado += len(filtroOfertasRecusado)
        numero_de_OfertasRetirado += len(filtroOfertasRetirado)
    limpar_tela()
    print("---------------------------------------------------------------------------------\nFinalidades de Consentimentos Universais (Classificados, Recusados e Retirados)\n---------------------------------------------------------------------------------")
    print("Email:             ", numero_de_EmailConcedido," | ", numero_de_EmailRecusado, " | ", numero_de_EmailRetirado)
    print("Notificações:      ", numero_de_NotificacoesConcedido," | ", numero_de_NotificacoesRecusado, " | ", numero_de_NotificacoesRetirado) 
    print("Ofertas do Cliente 2: ", numero_de_OfertasConcedido, " | ", numero_de_OfertasRecusado , " | ", numero_de_OfertasRetirado)
    print("SMS:               ", numero_de_SMSConcedido, " | ", numero_de_SMSRecusado, " | ", numero_de_SMSRetirado)
    print("Telefone:          ", numero_de_TelefoneConcedido, " | ", numero_de_TelefoneRecusado, " | ", numero_de_TelefoneRetirado)
    print("Whatsapp:          ", numero_de_WhatsappConcedido, " | ", numero_de_WhatsappRecusado, " | ", numero_de_WhatsappRetirado)

    ws[NotificacoesConcedidoCliente2] = numero_de_NotificacoesConcedido
    ws[NotificacoesRetiradoCliente2] = numero_de_NotificacoesRetirado
    ws[NotificacoesRecusadoCliente2] = numero_de_NotificacoesRecusado

    ws[SMSConcedidoCliente2] = numero_de_SMSConcedido
    ws[SMSRetiradoCliente2] = numero_de_SMSRetirado
    ws[SMSRecusadoCliente2] = numero_de_SMSRecusado

    ws[EmailConcedidoCliente2] = numero_de_EmailConcedido
    ws[EmailRetiradoCliente2] = numero_de_EmailRetirado
    ws[EmailRecusadoCliente2] = numero_de_EmailRecusado

    ws[TelefoneConcedidoCliente2] = numero_de_TelefoneConcedido
    ws[TelefoneRetiradoCliente2] = numero_de_TelefoneRetirado
    ws[TelefoneRecusadoCliente2] = numero_de_TelefoneRecusado

    ws[WhatsappConcedidoCliente2] = numero_de_WhatsappConcedido
    ws[WhatsappRetiradoCliente2] = numero_de_WhatsappRetirado
    ws[WhatsappRecusadoCliente2] = numero_de_WhatsappRecusado

    ws[OfertasConcedidoCliente2] = numero_de_OfertasConcedido
    ws[OfertasRetiradoCliente2] = numero_de_OfertasRetirado
    ws[OfertasRecusadoCliente2] = numero_de_OfertasRecusado

    # Salvar as alterações no arquivo Excel
    salvarExcel(graficoCliente,wb)
    time.sleep(3)
# Função para ler a Quantidade de Uso das APIs
def APIs(leituraAPIs, celulaTotalAPIs, graficoCliente, celulaTotalAPIsCliente2):
    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    totalAPIs = 0

    # Removendo o BOM da coluna e aspas duplas, se presentes
    leituraAPIs.columns = leituraAPIs.columns.str.replace('ï»¿', '', regex=False)  # Remover BOM
    leituraAPIs.columns = leituraAPIs.columns.str.replace('"', '', regex=False)  # Remover aspas duplas
    leituraAPIs.columns = leituraAPIs.columns.str.strip()  # Remover espaços extras

    # Verificar as colunas após a limpeza
    print(leituraAPIs.columns)

    # Data atual
    data_atual = datetime.now()

    # Obtém o primeiro dia do mês atual
    primeiro_dia_mes_atual = data_atual.replace(day=1)

    # Obtém o primeiro dia do mês anterior
    primeiro_dia_mes_anterior = primeiro_dia_mes_atual - relativedelta(months=1)

    # Gera todas as datas do mês anterior
    dias_mes_anterior = pd.date_range(start=primeiro_dia_mes_anterior, end=primeiro_dia_mes_atual - pd.Timedelta(days=1))

    limpar_tela()

    # Verificar se a coluna '@timestamp per day' existe
    if '@timestamp per day' not in leituraAPIs.columns:
        print("'@timestamp per day' NÃO está presente no DataFrame.")
        return  # Sai da função se a coluna não for encontrada

    # Filtrando os dados para todas as datas do mês anterior
    filtro = leituraAPIs[leituraAPIs['@timestamp per day'].isin(dias_mes_anterior.strftime("%Y-%m-%d"))]

    # Converte para datetime, se necessário
    filtro['@timestamp per day'] = pd.to_datetime(filtro['@timestamp per day'], errors='coerce')

    # Mostrando os valores um por um
    for index, row in filtro.iterrows():
        # Verificar se as colunas existem
        if '@timestamp per day' in row and 'Count' in row:
            try:
                count_value = row['Count'].replace(',', '')  # Remove vírgulas
                count_int = int(count_value)  # Converte para inteiro
                print(f"Data: {row['@timestamp per day'].strftime('%d/%m/%Y')}, Count: {count_int}")
                
                ws[f'B{celulaTotalAPIsCliente2}'] = row['@timestamp per day'].strftime("%d/%m/%Y")
                ws[f'C{celulaTotalAPIsCliente2}'] = count_int
                totalAPIs += count_int
                celulaTotalAPIsCliente2 += 1
            except Exception as e:
                print(f"Erro ao processar a linha {index}: {e}")
        else:
            print(f"Coluna não encontrada na linha {index}.")

    print("------------------------------\nTotal de Uso das APIs: ", totalAPIs)
    ws[celulaTotalAPIs] = totalAPIs

    salvarExcel(graficoCliente, wb)
    time.sleep(3)
#Função para Pegar a Finalidade por Status de Consentimento Universal Cliente3
def StatusxFinalidadeConsentimentoUniversalCliente3 (leituraConsentimentoUniversal,graficoCliente):
    
    numero_de_TransmissaoAutorizado=0; numero_de_TransmissaoRejeitado=0; numero_de_TransmissaoAguardando=0
    numero_de_RecepcaoAutorizado=0; numero_de_RecepcaoRejeitado=0; numero_de_RecepcaoAguardando=0
    somaAutorizados=0;somaRejeitados=0;somaAguardando=0

    ws, wb = CarregarExcel(graficoCliente)
    chunk_size = leitura()
    
    for chunk in leituraConsentimentoUniversal:
        filtroTransmissaoAutorizado = chunk[(chunk['Processing Purpose'] == "TransmissÃ£o de dados") & (chunk['Consent Purpose'] == "AUTHORIZED")]
        filtroTransmissaoRejeitado = chunk[(chunk['Processing Purpose'] == "TransmissÃ£o de dados") & (chunk['Consent Purpose'] == "REJECTED")]
        filtroTransmissaoAguardando = chunk[(chunk['Processing Purpose'] == "TransmissÃ£o de dados") & (chunk['Consent Purpose'] == "AWAITING AUTHORIZATION")]

        filtroRecepcaoAutorizado = chunk[(chunk['Processing Purpose'] == "RecepÃ§Ã£o de dados") & (chunk['Consent Purpose'] == "AUTHORIZED")]
        filtroRecepcaoRejeitado = chunk[(chunk['Processing Purpose'] == "RecepÃ§Ã£o de dados") & (chunk['Consent Purpose'] == "REJECTED")]
        filtroRecepcaoAguardando = chunk[(chunk['Processing Purpose'] == "RecepÃ§Ã£o de dados") & (chunk['Consent Purpose'] == "AWAITING AUTHORIZATION")]
 
        numero_de_TransmissaoAutorizado += len(filtroTransmissaoAutorizado)
        numero_de_TransmissaoRejeitado += len(filtroTransmissaoRejeitado)
        numero_de_TransmissaoAguardando += len(filtroTransmissaoAguardando)

        numero_de_RecepcaoAutorizado += len(filtroRecepcaoAutorizado)
        numero_de_RecepcaoRejeitado += len(filtroRecepcaoRejeitado)
        numero_de_RecepcaoAguardando += len(filtroRecepcaoAguardando)
        
        somaAutorizados = numero_de_TransmissaoAutorizado + numero_de_RecepcaoAutorizado
        somaRejeitados = numero_de_TransmissaoRejeitado + numero_de_RecepcaoRejeitado
        somaAguardando = numero_de_TransmissaoAguardando + numero_de_RecepcaoAguardando
    limpar_tela()
    print("---------------------------------------------------------------------------------------------------\nFinalidades de Consentimentos Universais (Autorizado, Rejeitado, Aguardando Autorização)\n---------------------------------------------------------------------------------------------------")
    print("Transmissão de dados: ", numero_de_TransmissaoAutorizado," | ", numero_de_TransmissaoRejeitado, " | ", numero_de_TransmissaoAguardando)
    print("Recepção de dados: ", numero_de_RecepcaoAutorizado," | ", numero_de_RecepcaoRejeitado, " | ", numero_de_RecepcaoAguardando) 
    print("------Total------\nAutorizados:", somaAutorizados, " | Rejeitados:", somaRejeitados," | Aguardando Autorização: ", somaAguardando)
    
    ws[TransmissaoAutorizadoCliente3] = numero_de_TransmissaoAutorizado
    ws[TransmissaoRejeitadoCliente3] = numero_de_TransmissaoRejeitado
    ws[TransmissaoAguardandoCliente3] = numero_de_TransmissaoAguardando

    ws[RecepcaoAutorizadoCliente3] = numero_de_RecepcaoAutorizado
    ws[RecepcaoRejeitadoCliente3] = numero_de_RecepcaoRejeitado
    ws[RecepcaoAguardandoCliente3] = numero_de_RecepcaoAguardando
    
    ws[somaAutorizadosCliente3] = somaAutorizados
    ws[somaRejeitadosCliente3] = somaRejeitados
    ws[somaAguardandoCliente3] = somaAguardando
    
    ws[somaAutorizadosHistoricoCliente3] = somaAutorizados
    ws[somaRejeitadosHistoricoCliente3] = somaRejeitados
    ws[somaAguardandoHistoricoCliente3] = somaAguardando
    
    # Salvar as alterações no arquivo Excel
    salvarExcel(graficoCliente,wb)
    time.sleep(3)
print (r"""
...............................................................................
.   ________     ________    __    __         __   ________   ___    ___      .
.  |   __   |   /  ______|  |  |  |  |       |  | |__    __|  \  \  /  /      .
.  |  |__|  |  |  |   ___   |  |  |  |       |  |    |  |      \  \/  /       .
.  |   __   |  |  |  |_  \  |  |  |  |       |  |    |  |       \    /        .
.  |  |  |  |  |  |____| |  |  |  |  |____   |  |    |  |        |  |         .
.  |__|  |__|   \________/  |__|  |_______|  |__|    |__|        |__|  REPORT .
...............................................................................
""")
time.sleep(1)
print ("Seja bem vindo!!!")
time.sleep(1)
while True: 
    # Início do contador de tempo
    inicio = time.time()
    limpar_tela()
    opcao = input ("Selecione uma opção\n1- Cliente 1\n2- Cliente 2\n3- Cliente 3\n4- Sair do programa\nEscolha sua opção: ")
    opcao = opcao.lower()
    if opcao == "sair" or opcao == "4":

        limpar_tela()
        print("Programa encerrado!!!")
        break

    while True: 
        limpar_tela()
        try: 
            caminho = input("Passe o caminho para a pasta do relatório: ")
            os.chdir(caminho)
            print("\n\033[1mDiretório atual:\033[0m", os.getcwd())
            time.sleep(2)
            break
        except FileNotFoundError:
            print(f"O diretório '{caminho}' não foi encontrado. Certifique-se de que o nome do caminho esteja correto!")
            limpar_tela()
    if opcao == "cliente1" or opcao == "1":
        while True:
            limpar_tela()
            print("---------------- Cliente 1 ----------------")
            opcaoFiltragem = input("Selecione uma opção\n"
                    "1- Filtragem de todos os dados\n"
                    "2- Filtragem das DSRs\n"
                    "3- Filtragem do Histórico de Cookies\n"
                    "4- Filtragem dos Sistemas Operacionais de Consentimento de Cookies\n"
                    "5- Filtragem das Origens dos Consentimentos de Cookies\n"
                    "6- Filtragem da Classificação de Cookies\n"
                    "7- Voltar\n"
                    "Escolha sua opção: ")

            if opcaoFiltragem == "1":
                chunk_size=leitura() 
                leituraDSR = leituraArquivosDSRs()
                limpar_tela()
                leituraCookies, arquivo_Cookies = leituraArquivosCookies()
                FiltragemDSRs(leituraDSR,ExcelCliente1,opcao,linhainicialDSRsCliente1,celulaTotalCliente1,celulaNaoVenda)
                limpar_tela()
                HistoricoCookies(leituraCookies,ExcelCliente1,opcao,celulaStatusConcedidosCliente1,celulaStatusIgnoradosCliente1,
                 celulaStatusRetiradosCliente1,celulaStatusRecusadosCliente1,celulaConcedidosCliente1,celulaRetiradosCliente1,
                 celulaIgnoradosCliente1,celulaRecusadosCliente1)
                limpar_tela()   
                leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                SistemasOperacionaisCookies(leituraCookies, linhainicialSoSCliente1, ExcelCliente1)
                limpar_tela()   
                leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                OrigemConsentimentoCookiesCliente1(leituraCookies,ExcelCliente1,origensCliente1)
                limpar_tela()    
                leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                ClassificacaoCookies(leituraCookies, ExcelCliente1, ConcedidoEssencialCliente1, RecusadoEssencialCliente1, 
                            RetiradoEssencialCliente1, ConcedidoPublicidadeCliente1, RecusadoPublicidadeCliente1, 
                            RetiradoPublicidadeCliente1, ConcedidoAnaliseEPersonalizacaoCliente1, 
                            RecusadoAnaliseEPersonalizacaoCliente1, RetiradoAnaliseEPersonalizacaoCliente1, 
                            ConcedidoDesempenhoEFuncionalidadeCliente1, RecusadoDesempenhoEFuncionalidadeCliente1, 
                            RetiradoDesempenhoEFuncionalidadeCliente1)
                
                tempoDecorrido(inicio)
                    
            elif opcaoFiltragem == "2":

                leituraDSR = leituraArquivosDSRs()
                FiltragemDSRs(leituraDSR,ExcelCliente1,opcao,linhainicialDSRsCliente1,celulaTotalCliente1, celulaNaoVenda)
                    
            elif opcaoFiltragem == "3":
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                HistoricoCookies(leituraCookies,ExcelCliente1,opcao,celulaStatusConcedidosCliente1,celulaStatusIgnoradosCliente1,
                 celulaStatusRetiradosCliente1,celulaStatusRecusadosCliente1,celulaConcedidosCliente1,celulaRetiradosCliente1,
                 celulaIgnoradosCliente1,celulaRecusadosCliente1)
                
            elif opcaoFiltragem =="4":
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                SistemasOperacionaisCookies(leituraCookies, linhainicialSoSCliente1, ExcelCliente1)

            elif opcaoFiltragem =="5":
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                OrigemConsentimentoCookiesCliente1(leituraCookies,ExcelCliente1,origensCliente1)

            elif opcaoFiltragem =="6": 
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                ClassificacaoCookies(leituraCookies, ExcelCliente1, ConcedidoEssencialCliente1, RecusadoEssencialCliente1, 
                            RetiradoEssencialCliente1, ConcedidoPublicidadeCliente1, RecusadoPublicidadeCliente1, 
                            RetiradoPublicidadeCliente1, ConcedidoAnaliseEPersonalizacaoCliente1, 
                            RecusadoAnaliseEPersonalizacaoCliente1, RetiradoAnaliseEPersonalizacaoCliente1, 
                            ConcedidoDesempenhoEFuncionalidadeCliente1, RecusadoDesempenhoEFuncionalidadeCliente1, 
                            RetiradoDesempenhoEFuncionalidadeCliente1)

            elif opcaoFiltragem =="7" or opcaoFiltragem=="voltar":

                print("Retornando!")
                break

            else:
                print ("Opção inválida!!!")
                time.sleep (2)
    elif opcao == "cliente2" or opcao == "2":
        while True:
            limpar_tela()
            print("---------------- Cliente 2 ----------------")
            opcaoFiltragem = input("Selecione uma opção:\n"
                    "1- Filtragem de todos os dados\n"
                    "2- Filtragem das DSRs\n"
                    "3- Filtragem da Origem dos Consentimentos Universais\n"
                    "4- Filtragem do Histórico de Consentimentos Universais\n"
                    "5- Filtragem da Finalidade de Consentimentos Universais\n"
                    "6- Filtragem do Histórico de Cookies\n"
                    "7- Filtragem do Histórico de Cookies Cliente 2 Segundo Ambiente\n"
                    "8- Filtragem dos Sistemas Operacionais de Consentimento de Cookies\n"
                    "9- Filtragem das Origens dos Consentimentos de Cookies\n"
                    "10- Filtragem da Classificação de Cookies\n"
                    "11- Filtragem dos Acessos de APIs\n"
                    "12- Voltar\n"
                    "Escolha uma opção: ")
            if opcaoFiltragem == "1":
                limpar_tela()
                chunk_size=leitura()

                leituraDSR = leituraArquivosDSRs()
                leituraConsentimentoUniversal, arquivoConsentimentoUniversal = leituraArquivosConsentimentoUniversal()
                limpar_tela()
                print ("-------- Cliente 2 --------")
                leituraCookies, arquivo_Cookies = leituraArquivosCookies()
                limpar_tela()
                leituraCookiesCliente2Ambiente,arquivo_CookiesCliente2Ambiente = leituraArquivosCookiesCliente2Ambiente()
                leituraAPIs = leituraArquivoAPIs()

                FiltragemDSRs(leituraDSR,ExcelCliente2,opcao,linhainicialDSRsCliente2,celulaTotalCliente2, celulaNaoVenda)
                   
                OrigemConsentimentoUniversalCliente2(leituraConsentimentoUniversal,ExcelCliente2)

                leituraConsentimentoUniversal = pd.read_csv(arquivoConsentimentoUniversal+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                HistoricoConsentimentoUniversal(leituraConsentimentoUniversal,ExcelCliente2,ConsentimentoUniversalConcedidoCliente2,
                                                ConsentimentoUniversalRetiradoCliente2, ConsentimentoUniversalRecusadoCliente2)

                leituraConsentimentoUniversal = pd.read_csv(arquivoConsentimentoUniversal+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                FinalidadeConsentimentoUniversalCliente2 (leituraConsentimentoUniversal,ExcelCliente2)

                HistoricoCookies(leituraCookies,ExcelCliente2,opcao,celulaStatusConcedidosCliente2,celulaStatusIgnoradosCliente2,
                 celulaStatusRetiradosCliente2,celulaStatusRecusadosCliente2,celulaConcedidosCliente2,celulaRetiradosCliente2,
                 celulaIgnoradosCliente2,celulaRecusadosCliente2)
 
                HistoricoCookiesCliente2Ambiente(leituraCookiesCliente2Ambiente,ExcelCliente2)

                leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                SistemasOperacionaisCookies(leituraCookies, linhainicialSoSCliente2, ExcelCliente2)
                    
                leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                leituraCookiesEye = pd.read_csv(arquivo_CookiesCliente2Ambiente+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                OrigemConsentimentoCookiesCliente2(leituraCookies, leituraCookiesCliente2Ambiente, origensCliente2, ExcelCliente2)

                leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                ClassificacaoCookies(leituraCookies, ExcelCliente2, ConcedidoEssencialCliente2, RecusadoEssencialCliente2, 
                            RetiradoEssencialCliente2, ConcedidoPublicidadeCliente2, RecusadoPublicidadeCliente2, 
                            RetiradoPublicidadeCliente2, ConcedidoAnaliseEPersonalizacaoCliente2, 
                            RecusadoAnaliseEPersonalizacaoCliente2, RetiradoAnaliseEPersonalizacaoCliente2, 
                            ConcedidoDesempenhoEFuncionalidadeCliente2, RecusadoDesempenhoEFuncionalidadeCliente2, 
                            RetiradoDesempenhoEFuncionalidadeCliente2)

                APIs(leituraAPIs, celulaTotalAPIs,ExcelCliente2,celulaTotalAPIsCliente2)
                
                tempoDecorrido(inicio)
            elif opcaoFiltragem == "2":

                leituraDSR = leituraArquivosDSRs()
                FiltragemDSRs(leituraDSR,ExcelCliente2,opcao,linhainicialDSRsCliente2,celulaTotalCliente2, celulaNaoVenda)
            elif opcaoFiltragem == "3":

                leituraConsentimentoUniversal, _ = leituraArquivosConsentimentoUniversal()
                OrigemConsentimentoUniversalCliente2(leituraConsentimentoUniversal,ExcelCliente2)                
            elif opcaoFiltragem == "4":

                leituraConsentimentoUniversal, _ = leituraArquivosConsentimentoUniversal()
                HistoricoConsentimentoUniversal(leituraConsentimentoUniversal,ExcelCliente2,ConsentimentoUniversalConcedidoCliente2,
                                                ConsentimentoUniversalRetiradoCliente2, ConsentimentoUniversalRecusadoCliente2)
            elif opcaoFiltragem == "5":
                leituraConsentimentoUniversal, _ = leituraArquivosConsentimentoUniversal()
                FinalidadeConsentimentoUniversalCliente2 (leituraConsentimentoUniversal,ExcelCliente2)
            elif opcaoFiltragem == "6":
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                HistoricoCookies(leituraCookies,ExcelCliente2,opcao,celulaStatusConcedidosCliente2,celulaStatusIgnoradosCliente2,
                 celulaStatusRetiradosCliente2,celulaStatusRecusadosCliente2,celulaConcedidosCliente2,celulaRetiradosCliente2,
                 celulaIgnoradosCliente2,celulaRecusadosCliente2)
            elif opcaoFiltragem == "7":
                limpar_tela()
                leituraCookiesCliente2Ambiente, _ = leituraArquivosCookiesCliente2Ambiente()
                HistoricoCookiesCliente2Ambiente(leituraCookiesEye,ExcelCliente2)
            elif opcaoFiltragem == "8":
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                SistemasOperacionaisCookies(leituraCookies, linhainicialSoSCliente2, ExcelCliente2)                
            elif opcaoFiltragem == "9":
                limpar_tela()
                print ("--------- Cliente2 ---------")
                leituraCookies, _ = leituraArquivosCookies()
                leituraCookiesCliente2Ambiente, _ = leituraArquivosCookiesCliente2Ambiente()
                OrigemConsentimentoCookiesCliente2(leituraCookies, leituraCookiesCliente2Ambiente, origensCliente2, ExcelCliente2)
            elif opcaoFiltragem == "10":
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                ClassificacaoCookies(leituraCookies, ExcelCliente2, ConcedidoEssencialCliente2, RecusadoEssencialCliente2, 
                            RetiradoEssencialCliente2, ConcedidoPublicidadeCliente2, RecusadoPublicidadeCliente2, 
                            RetiradoPublicidadeCliente2, ConcedidoAnaliseEPersonalizacaoCliente2, 
                            RecusadoAnaliseEPersonalizacaoCliente2, RetiradoAnaliseEPersonalizacaoCliente2, 
                            ConcedidoDesempenhoEFuncionalidadeCliente2, RecusadoDesempenhoEFuncionalidadeCliente2, 
                            RetiradoDesempenhoEFuncionalidadeCliente2)
            elif opcaoFiltragem =="11":
                limpar_tela()
                leituraAPIs = leituraArquivoAPIs()
                APIs(leituraAPIs, celulaTotalAPIs,ExcelCliente2,celulaTotalAPIsCliente2)
            elif opcaoFiltragem == "12":
                print("Retornando!")
                break
            else:
                print ("Opção inválida!!!")
                time.sleep (2)
    elif opcao == "cliente3" or opcao == "3":

        while True:
            limpar_tela()
            print("---------------- Cliente 3 ----------------")
            opcaoFiltragem = input("Selecione uma opção:\n"
                    "1- Filtragem de todos os dados\n"
                    "2- Filtragem do Status de Consentimentos Universais\n"
                    "3- Filtragem da Finalidade de Consentimentos Universais\n"
                    "4- Filtragem do Histórico de Cookies\n"
                    "5- Filtragem dos Sistemas Operacionais de Consentimento de Cookies\n"
                    "6- Filtragem das Origens dos Consentimentos de Cookies\n"
                    "7- Filtragem da Classificação de Cookies\n"
                    "8- Voltar\n"
                    "Escolha uma opção: ")
            if opcaoFiltragem == "1":
                chunk_size=leitura()                
                leituraConsentimentoUniversal, arquivoConsentimentoUniversal = leituraArquivosConsentimentoUniversal()
                leituraCookies, arquivo_Cookies = leituraArquivosCookies()

                HistoricoConsentimentoUniversal(leituraConsentimentoUniversal,ExcelCliente3,ConsentimentoUniversalConcedidoCliente3,
                                                ConsentimentoUniversalRetiradoCliente3,
                                                ConsentimentoUniversalRecusadoCliente3)

                leituraConsentimentoUniversal = pd.read_csv(arquivoConsentimentoUniversal+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                StatusxFinalidadeConsentimentoUniversalCliente3(leituraConsentimentoUniversal,ExcelCliente3)

                HistoricoCookies(leituraCookies,ExcelCliente3,opcao,celulaStatusConcedidosCliente3,celulaStatusIgnoradosCliente3,
                 celulaStatusRetiradosCliente3,celulaStatusRecusadosCliente3,celulaConcedidosCliente3,celulaRetiradosCliente3,
                 celulaIgnoradosCliente3,celulaRecusadosCliente3)

                leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                SistemasOperacionaisCookies(leituraCookies, linhainicialSoSCliente3, ExcelCliente3)
                    
                leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                OrigemConsentimentoCookiesCliente3(leituraCookies,ExcelCliente3,origensCliente3)

                leituraCookies = pd.read_csv(arquivo_Cookies+'.csv', encoding="latin-1", sep=';', chunksize=chunk_size)
                ClassificacaoCookies(leituraCookies, ExcelCliente3, ConcedidoEssencialCliente3, RecusadoEssencialCliente3, 
                            RetiradoEssencialCliente3, ConcedidoPublicidadeCliente3, RecusadoPublicidadeCliente3, 
                            RetiradoPublicidadeCliente3, ConcedidoAnaliseEPersonalizacaoCliente3, 
                            RecusadoAnaliseEPersonalizacaoCliente3, RetiradoAnaliseEPersonalizacaoCliente3, 
                            ConcedidoDesempenhoEFuncionalidadeCliente3, RecusadoDesempenhoEFuncionalidadeCliente3, 
                            RetiradoDesempenhoEFuncionalidadeCliente3)
                
                tempoDecorrido(inicio)
                
            elif opcaoFiltragem == "2":
                leituraConsentimentoUniversal, _ = leituraArquivosConsentimentoUniversal()
                HistoricoConsentimentoUniversal(leituraConsentimentoUniversal,ExcelCliente3,ConsentimentoUniversalConcedidoCliente3,
                                                ConsentimentoUniversalRetiradoCliente3, ConsentimentoUniversalRecusadoCliente3)
            elif opcaoFiltragem == "3":
                leituraConsentimentoUniversal, _ = leituraArquivosConsentimentoUniversal()
                StatusxFinalidadeConsentimentoUniversalCliente3(leituraConsentimentoUniversal,ExcelCliente3)
            elif opcaoFiltragem == "4":
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                HistoricoCookies(leituraCookies,ExcelCliente3,opcao,celulaStatusConcedidosCliente3,celulaStatusIgnoradosCliente3,
                 celulaStatusRetiradosCliente3,celulaStatusRecusadosCliente3,celulaConcedidosCliente3,celulaRetiradosCliente3,
                 celulaIgnoradosCliente3,celulaRecusadosCliente3)
            elif opcaoFiltragem == "5":
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                SistemasOperacionaisCookies(leituraCookies, linhainicialSoSCliente3, ExcelCliente3)   
            elif opcaoFiltragem == "6":  
                limpar_tela()                  
                leituraCookies, _ = leituraArquivosCookies()
                OrigemConsentimentoCookiesCliente3(leituraCookies,ExcelCliente3,origensCliente3)
            elif opcaoFiltragem == "7":
                limpar_tela()
                leituraCookies, _ = leituraArquivosCookies()
                ClassificacaoCookies(leituraCookies, ExcelCliente3, ConcedidoEssencialCliente3, RecusadoEssencialCliente3, 
                            RetiradoEssencialCliente3, ConcedidoPublicidadeCliente3, RecusadoPublicidadeCliente3, 
                            RetiradoPublicidadeCliente3, ConcedidoAnaliseEPersonalizacaoCliente3, 
                            RecusadoAnaliseEPersonalizacaoCliente3, RetiradoAnaliseEPersonalizacaoCliente3, 
                            ConcedidoDesempenhoEFuncionalidadeCliente3, RecusadoDesempenhoEFuncionalidadeCliente3, 
                            RetiradoDesempenhoEFuncionalidadeCliente3)
            elif opcaoFiltragem == "8":
                print("Retornando!")
                break
            else:
                print ("Opção inválida!!!")
                time.sleep (2)
    else:
        print("Opção inválida")
